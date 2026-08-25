"""独立知乎官方搜索 Provider；不导入其它平台代码。"""
from __future__ import annotations
import argparse, csv, json, os, sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen

API = "https://developer.zhihu.com/api/v1/content/zhihu_search"
FIELDS = ["平台","关键词","标题","内容类型","摘要","作者","作者主页","文章/回答链接","发布时间","排序分数","阅读量","点赞数","收藏数","评论数","抓取时间"]

def _parse_time(v):
    if not v: return None
    s = str(v).replace(" UTC", "+00:00").replace("Z", "+00:00")
    try: return datetime.fromisoformat(s)
    except ValueError: return None

def _get(obj, *keys):
    for k in keys:
        if obj.get(k) is not None: return obj[k]
    return ""

def search(query, secret):
    params = urlencode({"Query": query, "Count": 10})
    req = Request(f"{API}?{params}", headers={"Authorization": f"Bearer {secret}", "X-Request-Timestamp": str(int(datetime.now(timezone.utc).timestamp())), "Content-Type": "application/json"})
    try:
        with urlopen(req, timeout=30) as r: return json.loads(r.read().decode("utf-8"))
    except Exception as e:
        raise RuntimeError(f"知乎接口请求失败：{type(e).__name__}（未输出密钥或响应内容）") from e

def rows_for(query, payload, cutoff):
    data = payload.get("Data", payload.get("data", {})) if isinstance(payload, dict) else {}
    items = data.get("Items", data.get("items", [])) if isinstance(data, dict) else []
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    out=[]
    for x in items if isinstance(items, list) else []:
        t = _parse_time(_get(x,"EditTime","edit_time","CreatedTime","created_time","publish_time"))
        if t and t.astimezone() < cutoff: continue
        out.append({"平台":"知乎","关键词":query,"标题":_get(x,"Title","title"),"内容类型":_get(x,"ContentType","content_type","type"),"摘要":_get(x,"ContentText","content","excerpt","snippet"),"作者":_get(x,"AuthorName","author_name","author"),"作者主页":_get(x,"AuthorUrl","author_url"),"文章/回答链接":_get(x,"Url","url","link"),"发布时间":_get(x,"EditTime","edit_time","CreatedTime","created_time","publish_time"),"排序分数":_get(x,"RankingScore","ranking_score","score"),"阅读量":_get(x,"ReadCount","read_count","view_count"),"点赞数":_get(x,"VoteUpCount","voteup_count","like_count"),"收藏数":_get(x,"FavoriteCount","favorite_count","collect_count"),"评论数":_get(x,"CommentCount","comment_count"),"抓取时间":now})
    return out

def main():
    p=argparse.ArgumentParser(); p.add_argument("--query", nargs="+", required=True); p.add_argument("--months", type=int, default=6); p.add_argument("--top", type=int, default=10); p.add_argument("--output-dir", default="知乎输出")
    a=p.parse_args(); secret=os.getenv("ZHIHU_ACCESS_SECRET")
    if not secret: print("缺少 ZHIHU_ACCESS_SECRET；请先配置知乎 Access Secret。", file=sys.stderr); return 2
    cutoff=datetime.now().astimezone()-timedelta(days=30*a.months); all_rows=[]
    for q in a.query:
        try:
            qrows=rows_for(q, search(q,secret), cutoff)
            qrows.sort(key=lambda x: float(x["排序分数"] or 0), reverse=True)
            all_rows.extend(qrows[:a.top])
        
        except RuntimeError as e: print(str(e), file=sys.stderr); continue
    out=Path(a.output_dir); out.mkdir(parents=True, exist_ok=True)
    with (out/"知乎搜索.csv").open("w", newline="", encoding="utf-8-sig") as f:
        w=csv.DictWriter(f, fieldnames=FIELDS); w.writeheader(); w.writerows(all_rows)
    try:
        from openpyxl import Workbook
        wb=Workbook(); wb.remove(wb.active)
        for q in a.query:
            ws=wb.create_sheet((q[:28] or "关键词").replace("/","_").replace("\\","_"))
            ws.append(FIELDS)
            for row in [r for r in all_rows if r["关键词"] == q]: ws.append([row.get(k, "") for k in FIELDS])
            ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
            for cell in ws[1]: cell.font=cell.font.copy(bold=True)
        wb.save(out/"知乎搜索.xlsx")
    except ImportError:
        print("提示：未安装 openpyxl，仅输出 CSV；可执行 py -m pip install openpyxl", file=sys.stderr)
    print(f"完成：{len(all_rows)} 条，输出 {out/'知乎搜索.csv'} 与 {out/'知乎搜索.xlsx'}")
    return 0
if __name__ == "__main__": raise SystemExit(main())
