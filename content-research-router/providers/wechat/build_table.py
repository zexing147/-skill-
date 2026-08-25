from pathlib import Path
import argparse, re, yaml
from openpyxl import Workbook

ap=argparse.ArgumentParser(description='将公众号 Markdown 前言汇总为中文 Excel')
ap.add_argument('--input-dir', required=True)
ap.add_argument('--keyword', required=True)
ap.add_argument('--output', default='')
a=ap.parse_args()
ROOT=Path(a.input_dir).resolve()
OUT=Path(a.output).resolve() if a.output else ROOT/f'公众号_{a.keyword}_中文汇总.xlsx'
rows=[]
for p in ROOT.glob('*.md'):
    text=p.read_text(encoding='utf-8')
    fm=text.split('---',2)[1] if text.startswith('---') else ''
    data=yaml.safe_load(fm) or {}
    rows.append([data.get('title',''),data.get('author',''),data.get('date',''),data.get('source_url',''),data.get('description',''),data.get('captured_at',''),str(p)])
headers=['标题','公众号/作者','发布时间','原文链接','摘要','采集时间','本地正文文件']
wb=Workbook(); ws=wb.active; ws.title='文章信息'; ws.append(headers)
for r in rows: ws.append(r)
def add(name, data):
    s=wb.create_sheet(name); s.append(headers)
    for r in data: s.append(r)
dates=[]; unknown=[]
from datetime import datetime
for r in rows:
    try: dates.append((datetime.strptime(r[2].replace('年','-').replace('月','-').replace('日',''),'%Y-%m-%d %H:%M'),r))
    except Exception: unknown.append(r)
add('按发布时间排序',[r for _,r in sorted(dates,reverse=True)])
add('未判定日期',unknown)
log=wb.create_sheet('采集记录'); log.append(['关键词','候选上限','最终下载数','时间范围','说明']); log.append([a.keyword,50,len(rows),'近半年','按搜索结果顺序下载；公众号不抓评论；阅读量未可靠返回'])
for s in wb.worksheets:
    s.freeze_panes='A2'; s.auto_filter.ref=s.dimensions
    for c in s[1]: c.font=c.font.copy(bold=True)
    for col in s.columns: s.column_dimensions[col[0].column_letter].width=min(max(max(len(str(x.value or '')) for x in col)+2,12),45)
wb.save(OUT); print(OUT)
