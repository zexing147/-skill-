from __future__ import annotations
import argparse, re, shutil, subprocess, sys, urllib.request
from pathlib import Path

MC = Path(r"C:\Users\14709\Documents\ChatGPT\商业流程自动化\MediaCrawler")
PY = MC / ".venv" / "Scripts" / "python.exe"
BASE = MC / "config" / "base_config.py"
DY = MC / "config" / "dy_config.py"
MAP = {'aweme_id':'作品ID','aweme_type':'作品类型','title':'标题','desc':'作品描述','create_time':'发布时间','creator_hash':'创作者ID','nickname':'昵称','play_count':'播放量','liked_count':'点赞数','collected_count':'收藏数','comment_count':'评论数','share_count':'分享数','last_modify_ts':'最后修改时间','aweme_url':'作品链接','cover_url':'封面链接','video_download_url':'视频下载链接','music_download_url':'音乐下载链接','note_download_url':'图文下载链接','source_keyword':'来源关键词'}

def replace_list(text: str, name: str, value: str) -> str:
    return re.sub(rf"{name}\s*=\s*\[[\s\S]*?\]", f'{name} = [{value!r}]', text, count=1)

def translate_xlsx(src: Path, dst: Path) -> None:
    sys.path.insert(0, str(MC / '.venv' / 'Lib' / 'site-packages'))
    from openpyxl import load_workbook
    wb = load_workbook(src)
    for ws in wb.worksheets:
        for cell in ws[1]: cell.value = MAP.get(cell.value, cell.value)
        if ws.title == 'Contents':
            headers = [c.value for c in ws[1]]
            def sort_sheet(name, key):
                out = wb.create_sheet(name)
                out.append(headers)
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                if key not in headers:
                    out.append([f'当前接口未返回{key}，无法排序'] + [''] * (len(headers) - 1))
                    return
                idx = headers.index(key)
                rows.sort(key=lambda r: (r[idx] is not None, r[idx] or 0), reverse=True)
                for row in rows: out.append(list(row))
            sort_sheet('按播放量排序', '播放量')
            sort_sheet('按收藏量排序', '收藏数')
            sort_sheet('按评论数排序', '评论数')
    wb.save(dst)

def resolve_creator(value: str) -> str:
    if re.fullmatch(r'MS4wLjAB[A-Za-z0-9_-]+', value): return value
    req = urllib.request.Request(value, headers={'User-Agent':'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=30) as r: value = r.geturl()
    m = re.search(r'(?:/user/|[?&]sec_uid=)([^/?&#]+)', value)
    if not m: raise SystemExit('无法从主页链接解析 sec_user_id')
    return m.group(1)

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('mode', choices=('creator','detail'))
    ap.add_argument('--account'); ap.add_argument('--video')
    ap.add_argument('--videos', type=int, default=10)
    ap.add_argument('--comments', action=argparse.BooleanOptionalAction, default=True,
                    help='抓取一级评论；默认开启，使用 --no-comments 关闭')
    ap.add_argument('--output-dir', default='output')
    a = ap.parse_args()
    if not MC.exists() or not PY.exists(): raise SystemExit(f'MediaCrawler 环境不存在：{MC}')
    target = a.account if a.mode == 'creator' else a.video
    if not target: raise SystemExit(f'{a.mode} 模式必须提供对应输入')
    old_base, old_dy = BASE.read_text(encoding='utf-8'), DY.read_text(encoding='utf-8')
    try:
        base = old_base.replace('CRAWLER_MAX_NOTES_COUNT = 10', f'CRAWLER_MAX_NOTES_COUNT = {max(1,a.videos)}')
        base = re.sub(r'ENABLE_GET_COMMENTS\s*=\s*\w+', f'ENABLE_GET_COMMENTS = {bool(a.comments)}', base, count=1)
        BASE.write_text(base, encoding='utf-8')
        dy = old_dy
        if a.mode == 'creator': dy = replace_list(dy, 'DY_CREATOR_ID_LIST', resolve_creator(target))
        else: dy = replace_list(dy, 'DY_SPECIFIED_ID_LIST', target)
        DY.write_text(dy, encoding='utf-8')
        before = set((MC/'data'/'douyin').glob('*.xlsx'))
        env = dict(__import__('os').environ); env['PYTHONPATH'] = str(MC/'.venv'/'Lib'/'site-packages'); env['MPLCONFIGDIR'] = str(MC/'.matplotlib')
        cmd = [str(PY), 'main.py', '--platform', 'dy', '--lt', 'qrcode', '--type', 'creator' if a.mode=='creator' else 'detail', '--save_data_option', 'excel']
        result = subprocess.run(cmd, cwd=MC, env=env)
        if result.returncode: return result.returncode
        files = sorted(set((MC/'data'/'douyin').glob('*.xlsx')) - before, key=lambda p:p.stat().st_mtime, reverse=True)
        if not files: raise SystemExit('MediaCrawler 未生成 Excel')
        out = Path(a.output_dir).resolve(); out.mkdir(parents=True, exist_ok=True)
        translated = out / f'mediacrawler_{a.mode}_中文.xlsx'; translate_xlsx(files[0], translated)
        print(translated); return 0
    finally:
        BASE.write_text(old_base, encoding='utf-8'); DY.write_text(old_dy, encoding='utf-8')

if __name__ == '__main__': raise SystemExit(main())
